"""
Master pipeline: accepts PDFs from any supported bank (mixed together),
detects which bank each one is from, extracts transactions, categorizes
them, and generates a single consolidated Excel workbook with a
transactions view + summary by category + summary by month.
"""
import re
import shutil
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from parsers.liverpool import parse_liverpool
from parsers.banamex import parse_banamex
from parsers.invex import parse_invex
from parsers.nu import parse_nu
from parsers.banorte import parse_banorte
from parsers.santander import parse_santander
from categorize import categorize
from enrich import normalize_merchant, day_of_week_info
from msi_debt import active_msi, monthly_projection
import config
import db
import validate
import statements

BANK_DISPLAY = {
    "liverpool": "Liverpool", "banamex": "Banamex", "invex": "Invex Volaris",
    "nu": "Nu", "banorte": "Banorte", "santander": "Santander",
}

# NOTE: kept in Spanish because these are the literal month abbreviations
# printed on the statements themselves (used as dict keys for lookup).
SPANISH_MONTHS = {
    "ENE": 1, "FEB": 2, "MAR": 3, "ABR": 4, "MAY": 5, "JUN": 6,
    "JUL": 7, "AGO": 8, "SEP": 9, "OCT": 10, "NOV": 11, "DIC": 12,
}


def detect_bank(pdf_path: str) -> str:
    """Detects the bank first by filename, then falls back to page content."""
    name = Path(pdf_path).stem.upper()
    if "LIVERPOOL" in name:
        return "liverpool"
    if "BANAMEX" in name or "CITIBANAMEX" in name:
        return "banamex"
    if "INVEX" in name or "VOLARIS" in name:
        return "invex"
    if "BANORTE" in name:
        return "banorte"
    if "SANTANDER" in name:
        return "santander"
    # match NU only as a delimited token, so filenames like "NUMERO..." don't
    # false-positive as Nu
    if "NU" in re.split(r"[-_ ]", name):
        return "nu"

    # fallback: scan all pages (the bank logo is usually an image, so we
    # look for the company name that appears as plain text on inner pages).
    # Santander has no content-fallback branch: its PDFs have no extractable
    # text layer at all (image-only), so this scan would never see "SANTANDER"
    # without running OCR here - filename detection above is the only path.
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                text = page.extract_text() or ""
                if "LIVERPOOL" in text.upper():
                    return "liverpool"
                if "BANAMEX" in text.upper():
                    return "banamex"
                if "INVEX" in text.upper() or "VOLARIS" in text.upper():
                    return "invex"
                if "NU MÉXICO FINANCIERA" in text.upper() or "NU MEXICO FINANCIERA" in text.upper():
                    return "nu"
                if "BANORTE" in text.upper():
                    return "banorte"
            # last resort: if any page has Banamex's typical transaction
            # pattern (date date description +/- $amount), assume Banamex.
            # Checked after the BANORTE text match above since Banorte rows
            # share this exact date-date-description-amount shape.
            for page in pdf.pages:
                text = page.extract_text() or ""
                if any(re.match(
                    r"^\d{2}-\w{3}-\d{4}\s+\d{2}-\w{3}-\d{4}", line.strip(), re.IGNORECASE
                ) for line in text.split("\n")):
                    return "banamex"
    except Exception:
        pass
    return "unknown"


def liverpool_date_to_iso(date_str: str, year: int) -> str:
    day, mon = date_str.split("-")
    return f"{year:04d}-{SPANISH_MONTHS[mon]:02d}-{int(day):02d}"


def dashed_date_to_iso(date_str: str) -> str:
    # format "27-may-2026"
    day, mon, year = date_str.split("-")
    return f"{year}-{SPANISH_MONTHS[mon.upper()]:02d}-{int(day):02d}"


def guess_year_from_filename(pdf_path: str) -> int:
    # Only Liverpool needs this: its statement dates are "DD-MON" with no year
    # (the other banks' dates already carry YYYY). We take the 4-digit year from
    # the filename, falling back to the current year.
    # Note: a statement straddling a year boundary (Dec rows in a Jan
    # statement) will tag those rows with the statement's year. Include the
    # correct year in the filename to avoid it, or split by cut date if needed.
    fname = Path(pdf_path).stem.replace("_", "-")
    for token in fname.split("-"):
        if token.isdigit() and len(token) == 4:
            return int(token)
    return datetime.now().year


def build_rows(pdf_paths: list[str]) -> tuple[list[dict], list[dict], list[str]]:
    rows = []
    statement_rows = []
    imported_paths = []
    for pdf_path in pdf_paths:
        bank = detect_bank(pdf_path)
        fname = Path(pdf_path).stem
        ocr_text = None
        file_rows = []

        if bank == "liverpool":
            year = guess_year_from_filename(pdf_path)
            txns, _, ocr_text = parse_liverpool(pdf_path)
            for t in txns:
                cat = categorize(t["description"], t["amount"])
                file_rows.append({
                    "Bank": "Liverpool",
                    "File": fname,
                    "Date": liverpool_date_to_iso(t["date"], year),
                    "Description": t["description"],
                    "Category": cat,
                    "Amount": t["amount"],
                    "Type": t["type"],
                    "Review": "YES" if t["needs_review"] else "",
                })
        elif bank == "banamex":
            txns = parse_banamex(pdf_path)
            for t in txns:
                cat = categorize(t["description"], t["amount"])
                file_rows.append({
                    "Bank": "Banamex",
                    "File": fname,
                    "Date": dashed_date_to_iso(t["date"]),
                    "Description": t["description"],
                    "Category": cat,
                    "Amount": t["amount"],
                    "Type": t["type"],
                    "Review": "YES" if t.get("source") == "ocr_fallback" else "",
                    "ChargeDate": dashed_date_to_iso(t["charge_date"]) if t.get("charge_date") else None,
                    "Card": t.get("card"),
                })
        elif bank == "invex":
            txns = parse_invex(pdf_path)
            for t in txns:
                cat = categorize(t["description"], t["amount"])
                file_rows.append({
                    "Bank": "Invex Volaris",
                    "File": fname,
                    "Date": dashed_date_to_iso(t["date"]),
                    "Description": t["description"],
                    "Category": cat,
                    "Amount": t["amount"],
                    "Type": t["type"],
                    "Review": "YES" if t.get("is_installment") else "",
                    "is_installment": t.get("is_installment", False),
                    "ChargeDate": dashed_date_to_iso(t["charge_date"]) if t.get("charge_date") else None,
                    "StatementDate": dashed_date_to_iso(t["statement_date"]) if t.get("statement_date") else None,
                    "InstallmentNum": t.get("installment_num"),
                    "InstallmentTotal": t.get("installment_total"),
                    "OriginalAmount": t.get("original_amount"),
                    "RemainingBalance": t.get("remaining_balance"),
                    "Rate": t.get("rate"),
                })
        elif bank == "nu":
            txns = parse_nu(pdf_path)
            for t in txns:
                cat = categorize(t["description"], t["amount"])
                file_rows.append({
                    "Bank": "Nu",
                    "File": fname,
                    "Date": t["date"],
                    "Description": t["description"],
                    "Category": cat,
                    "Amount": t["amount"],
                    "Type": t["type"],
                    "Review": "",
                })
        elif bank == "banorte":
            txns = parse_banorte(pdf_path)
            for t in txns:
                cat = categorize(t["description"], t["amount"])
                file_rows.append({
                    "Bank": "Banorte",
                    "File": fname,
                    "Date": dashed_date_to_iso(t["date"]),
                    "Description": t["description"],
                    "Category": cat,
                    "Amount": t["amount"],
                    "Type": t["type"],
                    "Review": "",
                    "ChargeDate": dashed_date_to_iso(t["charge_date"]) if t.get("charge_date") else None,
                })
        elif bank == "santander":
            txns, ocr_text = parse_santander(pdf_path)
            for t in txns:
                cat = categorize(t["description"], t["amount"])
                file_rows.append({
                    "Bank": "Santander",
                    "File": fname,
                    "Date": dashed_date_to_iso(t["date"]),
                    "Description": t["description"],
                    "Category": cat,
                    "Amount": t["amount"],
                    "Type": t["type"],
                    "Review": "YES" if t["needs_review"] else "",
                })
        else:
            print(f"⚠️  Could not identify the bank for: {pdf_path} (skipped)")
            continue

        for r in file_rows:
            r["Month"] = r["Date"][:7]  # YYYY-MM
            r["DayOfWeek"], r["IsWeekend"] = day_of_week_info(r["Date"])
            r["MerchantNorm"] = normalize_merchant(r["Description"])

        ok, messages = validate.validate_file(bank, file_rows, pdf_path, ocr_text)
        for m in messages:
            print(f"  {fname}: {m}")
        if ok:
            rows.extend(file_rows)
            imported_paths.append(pdf_path)
            stmt = statements.extract_statement(bank, pdf_path, ocr_text)
            if stmt:
                statement_rows.append({
                    "Bank": BANK_DISPLAY[bank],
                    "Card": stmt["card"],
                    "PeriodStart": stmt["period_start"],
                    "PeriodEnd": stmt["period_end"],
                    "CutoffDate": stmt["cutoff_date"],
                    "PrevBalance": stmt["prev_balance"],
                    "ClosingBalance": stmt["closing_balance"],
                    "MinPayment": stmt["min_payment"],
                    "NoInterestPayment": stmt["no_interest_payment"],
                    "CreditLimit": stmt["credit_limit"],
                    "File": fname,
                })
        else:
            print(f"⛔ {fname}: failed validation, not imported")

    return rows, statement_rows, imported_paths


def write_excel(df: pd.DataFrame, statements_df: pd.DataFrame, out_path: str):
    wb = Workbook()

    # --- Sheet 1: Transactions ---
    ws = wb.active
    ws.title = "Transactions"
    headers = list(df.columns)
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF", name="Arial")
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.alignment = Alignment(horizontal="center")

    for _, row in df.iterrows():
        ws.append(list(row))

    amount_col_idx = headers.index("Amount") + 1
    for i, col in enumerate(headers, start=1):
        letter = get_column_letter(i)
        max_len = max([len(str(col))] + [len(str(v)) for v in df[col]]) if len(df) else len(col)
        ws.column_dimensions[letter].width = min(max_len + 2, 45)
        if col == "Amount":
            for r in range(2, len(df) + 2):
                ws[f"{letter}{r}"].number_format = "$#,##0.00;($#,##0.00)"
    ws.freeze_panes = "A2"

    n = len(df) + 1
    category_col = get_column_letter(headers.index("Category") + 1)
    month_col = get_column_letter(headers.index("Month") + 1) if "Month" in headers else None
    amount_col = get_column_letter(amount_col_idx)

    # --- Sheet 2: Summary by category ---
    ws2 = wb.create_sheet("Summary by Category")
    ws2.append(["Category", "Total", "# Transactions"])
    for cell in ws2[1]:
        cell.font = Font(bold=True, color="FFFFFF", name="Arial")
        cell.fill = PatternFill("solid", fgColor="4472C4")

    categories = sorted(df["Category"].unique()) if len(df) else []
    for i, cat in enumerate(categories, start=2):
        ws2.cell(row=i, column=1, value=cat)
        ws2.cell(row=i, column=2,
                  value=f"=SUMIFS(Transactions!{amount_col}2:{amount_col}{n},Transactions!{category_col}2:{category_col}{n},A{i})")
        ws2.cell(row=i, column=2).number_format = "$#,##0.00"
        ws2.cell(row=i, column=3,
                  value=f"=COUNTIF(Transactions!{category_col}2:{category_col}{n},A{i})")
    ws2.column_dimensions["A"].width = 32
    ws2.column_dimensions["B"].width = 16
    ws2.column_dimensions["C"].width = 16

    # --- Sheet 3: Summary by month ---
    if month_col:
        ws3 = wb.create_sheet("Summary by Month")
        ws3.append(["Month", "Total Spent (charges)", "Total Paid (payments)"])
        for cell in ws3[1]:
            cell.font = Font(bold=True, color="FFFFFF", name="Arial")
            cell.fill = PatternFill("solid", fgColor="4472C4")
        months = sorted(df["Month"].unique()) if len(df) else []
        type_col = get_column_letter(headers.index("Type") + 1)
        for i, month in enumerate(months, start=2):
            ws3.cell(row=i, column=1, value=month)
            ws3.cell(row=i, column=2,
                      value=(f'=SUMIFS(Transactions!{amount_col}2:{amount_col}{n},'
                             f'Transactions!{month_col}2:{month_col}{n},A{i},'
                             f'Transactions!{type_col}2:{type_col}{n},"charge")'))
            ws3.cell(row=i, column=2).number_format = "$#,##0.00"
            ws3.cell(row=i, column=3,
                      value=(f'=SUMIFS(Transactions!{amount_col}2:{amount_col}{n},'
                             f'Transactions!{month_col}2:{month_col}{n},A{i},'
                             f'Transactions!{type_col}2:{type_col}{n},"payment")'))
            ws3.cell(row=i, column=3).number_format = "$#,##0.00"
        ws3.column_dimensions["A"].width = 14
        ws3.column_dimensions["B"].width = 22
        ws3.column_dimensions["C"].width = 22

    # --- Sheet 4: Top Merchants (charges only) ---
    if "MerchantNorm" in headers and len(df):
        merchant_col = get_column_letter(headers.index("MerchantNorm") + 1)
        type_col = get_column_letter(headers.index("Type") + 1)
        ws4 = wb.create_sheet("Top Merchants")
        ws4.append(["Merchant", "Total Charged", "# Transactions"])
        for cell in ws4[1]:
            cell.font = Font(bold=True, color="FFFFFF", name="Arial")
            cell.fill = PatternFill("solid", fgColor="4472C4")
        # sort by actual spend so the highest merchants land on top; the
        # cells themselves stay live SUMIFS/COUNTIFS formulas.
        merchants = list(
            df[df["Type"] == "charge"].groupby("MerchantNorm")["Amount"]
            .sum().sort_values(ascending=False).index
        )
        for i, merch in enumerate(merchants, start=2):
            ws4.cell(row=i, column=1, value=merch)
            ws4.cell(row=i, column=2,
                      value=(f'=SUMIFS(Transactions!{amount_col}2:{amount_col}{n},'
                             f'Transactions!{merchant_col}2:{merchant_col}{n},A{i},'
                             f'Transactions!{type_col}2:{type_col}{n},"charge")'))
            ws4.cell(row=i, column=2).number_format = "$#,##0.00"
            ws4.cell(row=i, column=3,
                      value=(f'=COUNTIFS(Transactions!{merchant_col}2:{merchant_col}{n},A{i},'
                             f'Transactions!{type_col}2:{type_col}{n},"charge")'))
        ws4.column_dimensions["A"].width = 32
        ws4.column_dimensions["B"].width = 16
        ws4.column_dimensions["C"].width = 16

    # --- Sheet 5: Average ticket per category (charges only) ---
    if len(df):
        type_col = get_column_letter(headers.index("Type") + 1)
        ws5 = wb.create_sheet("Avg Ticket by Category")
        ws5.append(["Category", "Avg Charge", "# Charges"])
        for cell in ws5[1]:
            cell.font = Font(bold=True, color="FFFFFF", name="Arial")
            cell.fill = PatternFill("solid", fgColor="4472C4")
        for i, cat in enumerate(categories, start=2):
            ws5.cell(row=i, column=1, value=cat)
            ws5.cell(row=i, column=2,
                      value=(f'=AVERAGEIFS(Transactions!{amount_col}2:{amount_col}{n},'
                             f'Transactions!{category_col}2:{category_col}{n},A{i},'
                             f'Transactions!{type_col}2:{type_col}{n},"charge")'))
            ws5.cell(row=i, column=2).number_format = "$#,##0.00"
            ws5.cell(row=i, column=3,
                      value=(f'=COUNTIFS(Transactions!{category_col}2:{category_col}{n},A{i},'
                             f'Transactions!{type_col}2:{type_col}{n},"charge")'))
        ws5.column_dimensions["A"].width = 32
        ws5.column_dimensions["B"].width = 16
        ws5.column_dimensions["C"].width = 16

    # --- Sheet 6: Committed MSI debt (Phase 4.3) - precomputed, not SUMIFS:
    # it's a forward projection over future months that have no transaction
    # rows yet, so there's nothing for a live formula to sum. ---
    if "InstallmentTotal" in headers and len(df):
        active = active_msi(df)
        if len(active):
            ws6 = wb.create_sheet("Committed MSI Debt")
            ws6.append(["Bank", "Merchant", "Monthly Amount", "Installment", "End Month"])
            for cell in ws6[1]:
                cell.font = Font(bold=True, color="FFFFFF", name="Arial")
                cell.fill = PatternFill("solid", fgColor="4472C4")
            row_i = 2
            for _, r in active.iterrows():
                ws6.cell(row=row_i, column=1, value=r["Bank"])
                ws6.cell(row=row_i, column=2, value=r["Description"])
                ws6.cell(row=row_i, column=3, value=r["Amount"])
                ws6.cell(row=row_i, column=3).number_format = "$#,##0.00"
                ws6.cell(row=row_i, column=4, value=f'{int(r["InstallmentNum"])}/{int(r["InstallmentTotal"])}')
                ws6.cell(row=row_i, column=5, value=r["EndMonth"])
                row_i += 1
            ws6.column_dimensions["A"].width = 14
            ws6.column_dimensions["B"].width = 32
            ws6.column_dimensions["C"].width = 16
            ws6.column_dimensions["D"].width = 12
            ws6.column_dimensions["E"].width = 12

            projection = monthly_projection(active)
            if projection:
                row_i += 1
                ws6.cell(row=row_i, column=1, value="Month").font = Font(bold=True)
                ws6.cell(row=row_i, column=2, value="Total Committed").font = Font(bold=True)
                row_i += 1
                for month, total in projection:
                    ws6.cell(row=row_i, column=1, value=month)
                    ws6.cell(row=row_i, column=2, value=total)
                    ws6.cell(row=row_i, column=2).number_format = "$#,##0.00"
                    row_i += 1

    # --- Sheet 7: Credit Utilization (Phase 5) - precomputed, one row per
    # statement; only banks with a printed credit limit show up (Santander
    # is debit, Liverpool doesn't print a limit). ---
    util_df = statements_df[statements_df["CreditLimit"].notna()] if len(statements_df) else statements_df
    if len(util_df):
        ws7 = wb.create_sheet("Credit Utilization")
        ws7.append(["Bank", "Cutoff Date", "Closing Balance", "Credit Limit", "Utilization %"])
        for cell in ws7[1]:
            cell.font = Font(bold=True, color="FFFFFF", name="Arial")
            cell.fill = PatternFill("solid", fgColor="4472C4")
        for i, (_, r) in enumerate(util_df.sort_values("CutoffDate").iterrows(), start=2):
            ws7.cell(row=i, column=1, value=r["Bank"])
            ws7.cell(row=i, column=2, value=r["CutoffDate"])
            ws7.cell(row=i, column=3, value=r["ClosingBalance"])
            ws7.cell(row=i, column=3).number_format = "$#,##0.00"
            ws7.cell(row=i, column=4, value=r["CreditLimit"])
            ws7.cell(row=i, column=4).number_format = "$#,##0.00"
            if r["ClosingBalance"] is not None and r["CreditLimit"]:
                ws7.cell(row=i, column=5, value=r["ClosingBalance"] / r["CreditLimit"])
                ws7.cell(row=i, column=5).number_format = "0.0%"
        ws7.column_dimensions["A"].width = 14
        ws7.column_dimensions["B"].width = 14
        ws7.column_dimensions["C"].width = 16
        ws7.column_dimensions["D"].width = 16
        ws7.column_dimensions["E"].width = 14

    # --- Sheet 8: Fees & Interest by Year - categorize.py already routes
    # COMISION/INTERES/ANUALIDAD/IVA rows to "Card Interest/Fees" (Phase 1);
    # this is just a yearly rollup of that existing category, live SUMIFS. ---
    if "Category" in headers and len(df):
        ws8 = wb.create_sheet("Fees & Interest by Year")
        ws8.append(["Year", "Total Fees & Interest"])
        for cell in ws8[1]:
            cell.font = Font(bold=True, color="FFFFFF", name="Arial")
            cell.fill = PatternFill("solid", fgColor="4472C4")
        years = sorted({m[:4] for m in df["Month"]})
        month_col = get_column_letter(headers.index("Month") + 1)
        for i, year in enumerate(years, start=2):
            ws8.cell(row=i, column=1, value=year)
            ws8.cell(row=i, column=2,
                      value=(f'=SUMIFS(Transactions!{amount_col}2:{amount_col}{n},'
                             f'Transactions!{category_col}2:{category_col}{n},"Card Interest/Fees",'
                             f'Transactions!{month_col}2:{month_col}{n},">="&"{year}-01",'
                             f'Transactions!{month_col}2:{month_col}{n},"<="&"{year}-12")'))
            ws8.cell(row=i, column=2).number_format = "$#,##0.00"
        ws8.column_dimensions["A"].width = 10
        ws8.column_dimensions["B"].width = 20

    wb.save(out_path)


if __name__ == "__main__":
    pdf_paths = sys.argv[1:]
    if not pdf_paths:
        # no arguments: process everything in statements/
        folder = config.STATEMENTS_DIR
        pdf_paths = [str(p) for p in sorted(folder.glob("*.pdf"))]
        if not pdf_paths:
            print(f"No PDFs found in {folder}/ and no file was passed as an argument.")
            print("Usage: python3 build_report.py [file1.pdf file2.pdf ...]")
            sys.exit(1)
        print(f"Processing {len(pdf_paths)} PDF(s) from {folder}/ ...")

    rows, statement_rows, imported_paths = build_rows(pdf_paths)
    conn = db.connect()
    db.init_db(conn)
    n_new = db.insert_transactions(conn, rows)

    for stmt in statement_rows:
        warning = validate.check_continuity(conn, stmt)
        if warning:
            print(f"  ⚠️  {warning}")
    n_stmt_new, stmt_duplicates = db.insert_statements(conn, statement_rows)
    for dup in stmt_duplicates:
        print(f"  ⚠️  {dup['Bank']} {dup.get('CutoffDate')}: statement already imported, skipped")

    df = db.fetch_dataframe(conn)
    statements_df = db.fetch_statements(conn)

    out_dir = config.REPORTS_DIR
    out_dir.mkdir(exist_ok=True)
    out = str(out_dir / "consolidated_expense_report.xlsx")
    write_excel(df, statements_df, out)
    print(f"Imported {n_new} new of {len(rows)} parsed from {len(pdf_paths)} file(s).")
    print(f"Imported {n_stmt_new} new statement(s) ({len(stmt_duplicates)} duplicate(s) skipped).")
    print(f"Saved to {out}. DB now holds {len(df)} transactions total.")
    if len(df):
        print(df.groupby(["Bank", "Category"])["Amount"].sum().sort_values(ascending=False))

    # validated + inserted PDFs move out of statements/ so the next run
    # doesn't re-scan them; original file is kept (not deleted) for re-extraction
    # if a parser bug is found later - see ROADMAP.md 1.4.
    processed_dir = config.STATEMENTS_DIR / "processed"
    processed_dir.mkdir(parents=True, exist_ok=True)
    for pdf_path in imported_paths:
        src = Path(pdf_path)
        if src.resolve().parent == processed_dir.resolve():
            continue
        shutil.move(str(src), str(processed_dir / src.name))
    if imported_paths:
        print(f"Moved {len(imported_paths)} processed PDF(s) to {processed_dir}/")
